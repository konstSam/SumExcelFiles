import os
import locale
import openpyxl

# Set the locale to the user's default so to format the value in Euros/Dollars
locale.setlocale(locale.LC_ALL, '')

# Set the path to the parent directory
parent_dir = "C:\\path\\to\\parent\\directory\\"


def sum_excel_files(directory):  # Define a function to sum the values in the Excel files
    # Initialize the sum variable
    total_sum = 0
    unopened_files = []
    file_dict = {}

    # Iterate through the files and subdirectories in the directory
    for item in os.listdir(directory):
        # Construct the full path to the item
        item_path = os.path.join(directory, item)
        total_return = None

        # If the item is a directory, recursively call the function on the directory
        if os.path.isdir(item_path):
            sub_sum, sub_dict, sub_unopened_files = sum_excel_files(item_path)
            total_sum += sub_sum
            file_dict = {**file_dict, **sub_dict}
            unopened_files += sub_unopened_files

        # If the item is a file and ends with "Portfolio.xlsx"
        elif os.path.isfile(item_path) and item.endswith('Portfolio.xlsx'):
            # Load the workbook
            workbook = openpyxl.load_workbook(item_path, data_only=True)

            # Get the active sheet
            worksheet = workbook.active

            # Initialize the found variable to False
            found = False

            # Search for the cell with the value "Total Return"
            for row in worksheet.iter_rows(min_col=1, max_col=15):
                for cell in row:
                    if cell.value == 'Overview':
                        # Get the value of the cell below it
                        total_return = float(worksheet.cell(
                            row=cell.row + 3, column=cell.column).value)

                        # Print the value of Total Return and add it to the sum
                        # print(f'The value of Total Return in {directory}: {item} is: {total_return}')
                        found = True
                        total_sum += total_return
                        break

                if found:
                    break

            # Add to a list all those files that doesnt contain the "Overview" keyword
            if not found:
                unopened_files.append(item_path)

            # Close the workbook
            workbook.close()

        # Cache the return_total of every file for future use
        if total_return is not None:
            file_dict[item_path] = float(total_return)

    return total_sum, file_dict, unopened_files


# Define a function to iterate through a dictionary of all the opened files to get each ones return total
def fund_percentage(file_dict, total_sum):
    # Iterate through the files from the first time
    for file_path, return_total in file_dict.items():
        # Get the name of the file
        filename = os.path.basename(file_path)

        # Calculate the percentage of return_total with respect to total_sum
        percentage = (return_total / total_sum) * 100 if total_sum != 0 else 0
        print(
            f"File: {file_path}, Return Total: {locale.currency(return_total, grouping=True)} ({percentage:.2f}%)")


def main():
    try:
        # Call the function with the parent directory as the argument
        total_sum, file_dict, unopened_files = sum_excel_files(parent_dir)

        # Print the list of unopened files
        if unopened_files:
            print("The following files were opened but not found keyword 'Overview':")
            for file in unopened_files:
                print(f"X  File: {file}")
            print("\n")

        # Format the sum value
        formatted_sum = locale.currency(total_sum, grouping=True)

        # Call the function with 2 args from the `sum_excel_files`
        fund_percentage(file_dict, total_sum)

        # Print the value of all portfolios
        print(f"The value of all portfolios is: {formatted_sum}")

    except FileNotFoundError:
        print("The specified directory does not exist or cannot be found.")

    except PermissionError:
        print("The program does not have permission to access the specified directory or Excel files.")

    except ValueError:
        print("The cell value cannot be converted to a float value.")

    except openpyxl.utils.exceptions.CellCoordinatesException:
        print("The cell is not valid.")

    except Exception as e:
        print(f"An error occurred while processing portfolios: {str(e)}")


if __name__ == '__main__':
    main()
